#coding=utf-8
 
from xlwt3 import *
import urllib.request
import re
#import xlwt3 as xlwt
import string
from openpyxl import Workbook
from openpyxl.cell import get_column_letter


f = Font()
#f.height = 20*72
#f.name = 'Verdana'
#f.bold = True

f.underline = Font.UNDERLINE_SINGLE
f.colour_index = 4

h_style = XFStyle()
h_style.font = f






def getHtml(url):
    page = urllib.request.urlopen(url)
    html = page.read().decode('GBK')    
    return html


def getInfo(html,times):     
   

    if times ==1:
       infoall=[['区域','小区地段','房型','楼层','户型','装修',
              '面积','单价','价格','日期']]
    else:
        infoall=[]
    
    

    reg = r'<li class=lifw\d>([\s\S]+?)<li class=lixt>' #整条信息表达式（大段）   
    
	
	
    infore = re.compile(reg)
    infolist = re.findall(infore,html) #整条信息列表
    
    #print(len(infolist))

    reg1=r'target="_blank">(.+?)</a>'
    infore1 = re.compile(reg1)
    reg2=r'<a href="(\S+?)" target="_blank">'
    infore2 = re.compile(reg2)

    reg_dig = re.compile(r'(\d+(?:\.\d+)?)')
    
    

    
    for row in infolist:        
        inlowlist = re.findall(infore1,row) #单条信息具体内容
        for i in [6,7,8]:
            digial=re.search(reg_dig,inlowlist[i])
            if digial:                
                inlowlist[i]=float(digial.group(1))
        inlowlist.append("http://www.qidfc.cn"+re.search(infore2,row).group(1))#超级链接
        infoall.append(inlowlist)
    return infoall   
    
    
    # for info in infolist:#单条信息(未提取)
        # info_ax=[]#单条信息(提取)
        # col=0
        # for reg in reglist:
            # infore = re.compile(reg)
            # if col in [2,3,4]:               
                # #print(col)
                # #print(info_ax[0])
                # info_ax.append(float(re.search(infore,info).group(1)))
                
            # elif col == 6: 
                # info_ax+=getMore(re.search(infore,info).group(1))    
            # else:                    
                # #print(col)
                # #if len(info_ax)>0:
                   # # print(info_ax[0])
                # if re.search(infore,info)==None:
                    # col+=1
                    # info_ax.append('None')
                    # continue
                # info_ax.append(re.search(infore,info).group(1))
            # col+=1
        # infoall.append(info_ax)        
    # return infoall


def getMore(url):    
    html=getHtml(url)
    
    reg = r'<li class="des_cols2">(\S+)</li>'   
    infore = re.compile(reg)
    infolist=re.findall(infore,html) #多项信息列表

    reg=r'<p>(?:<p>)?(.+?)</p>'#详细信息
    infore = re.compile(reg)
    more=re.search(infore,html).group(1)

    
    infore = re.compile(r'(<.+?>)')
    more=infore.sub('', more,count = 10 )

    infore = re.compile(r'(&nbsp;)')
    more=infore.sub('', more,count = 10 )
    
    
    infolist+=[more,url]

    
    return infolist



def writexls(file_name,table_name,data_list):
    book=xlwt.Workbook()
    sheet=book.add_sheet(table_name)    
    row=0
   
    for datarow in data_list:
        
        
        col=0
        for data in datarow:
            #sheet.row(row).write(col,data)          
                
            if col==1 and row>0:
                sheet.write(row, col, Formula('HYPERLINK("%s"; "%s")'%(datarow[10],data)),h_style)                
            elif col==10:
                col+=1
            else:
                sheet.write(row, col, data)               
                
            #print(data)
            col+=1
        row+=1
        
    book.save(file_name)


def writexls_open(file_name,table_name,data_list):
    wb = Workbook()
    

    sheet = wb.active
    sheet.title = "range names"

    
    row=1
    for datarow in data_list:
        #ws.append(['%s' % data for data in datarow])
        col=1
        for data in datarow:
            col_idx = get_column_letter(col)
            sheet.cell('%s%s'%(col_idx, row)).value = data

            
            #sheet.Formula('HYPERLINK("%s"; "%s")'%(datarow[10],data))
            #sheet.formula('HYPERLINK("%s"; "%s")'%(datarow[10],data))
            #if col==1 and row>0:
            #    sheet.write(row, col, Formula('HYPERLINK("%s"; "%s")'%(datarow[10],data)),h_style)                
            #elif col==10:
            #    col+=1
            #else:
            #    sheet.write(row, col, data)
            col+=1
        row+=1
    sheet = wb.create_sheet()
    sheet.title = 'Pi'
    wb.save('%s.xlsx'%file_name)
    



# myinfo=[]
# for i in range(1,3):
    # html = getHtml("http://nt.58.com/qidong/ershoufang/e3/pn%d/"%i)
    # print("http://nt.58.com/qidong/ershoufang/e3/pn%d/"%i)
    # myinfo +=getInfo(html,i)

myinfo=[]
for i in range(1,51):
    html = getHtml("http://www.qidfc.cn/sale/page%s.html"%i)
    print("http://www.qidfc.cn/sale/page%s.html"%i)
    myinfo +=getInfo(html,i)
writexls_open('qidongC','test',myinfo)


